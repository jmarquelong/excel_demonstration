import openpyxl

# Prompts user to pay attention to case.
print('All inputs are case sensitive!')
# Prints blank line for formatting.
print(" ")

# Load file
file_name = str(input("Enter a file name: "))
file_type = '.xlsx'
concat = file_name + file_type
wb = openpyxl.load_workbook(concat)
ws = wb.active


# List for display_error function.
command_list = 'modify, review, titles, get, save,'
sheet_list = [wb.sheetnames]
# Prompts user for input.
print('Command list: modify, review, titles, get, save.\n'
      'Remember to save after finishing with your changes.')
print(" ")
user_input = input("Command: ")

# Functions follow:


def get_sheet():
    """User selects sheet to modify."""
    global ws
    sheet_input = input('Enter sheet name: ')
    ws = wb[sheet_input]
    print(ws)


def review_sheets():
    """Prints all sheet names"""
    print(wb.sheetnames)


def staff_contact():
    """Enter user input into cells using the 'modify' command."""
    phone = 10
    data = input('Enter phone number: ')
    if int(phone) == len(data):
        ws['D2'] = data
    elif len(data) != phone:
        display_error()


def save_doc():
    """Saves document changes."""
    wb.save(concat)


def worksheet_titles():
    """Loop through worksheets and rename each one."""
    for sheet in wb:
        first_cell_value = str(sheet['A2'].value)
        sheet.title = first_cell_value
        print(sheet.title)


def display_error():
    """Function can be used to minimize repeat coding."""
    print('Make a valid entry.')


# While loop takes user input and executes functions.

while user_input.lower() != 'q':
    if user_input == 'modify':
        staff_contact()
    elif user_input == 'review':
        review_sheets()
    elif user_input == 'titles':
        worksheet_titles()
    elif user_input == 'save':
        save_doc()
    elif user_input == 'get':
        get_sheet()
    elif user_input != command_list:
        display_error()

    user_input = input("Command: ")
print('Good bye!')